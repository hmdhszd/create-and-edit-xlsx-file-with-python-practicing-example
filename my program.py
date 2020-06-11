import openpyxl

workbook = openpyxl.load_workbook('example-XLSX-file.xlsx')

print(workbook.get_sheet_names())

sheet = workbook.get_sheet_by_name('Sheet1')


###########################

cell = sheet['B1']
print(cell.value)

print(sheet['B1'].value)


###########################

cell = sheet.cell(row = 2, column = 2)
print(cell.value)

print(sheet.cell(row = 2, column = 2).value)

###########################



###########################



wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')

sheet['A1'].value = "hamid"

wb.save('FileName.xlsx')



###########################

sheet2 = wb.create_sheet()
sheet2.title = "my new sheet"
wb.save('FileName2.xlsx')


